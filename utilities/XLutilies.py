import openpyxl
from openpyxl.styles import PatternFill

def GetRowsCount(file,sheetname):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	return sheet.max_row
def GetColumnsCount(file,sheetname):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	return sheet.max_column
def read(file,sheetname,rownum,columnnum):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	return sheet.cell(rownum,columnnum).value
def write(file,sheetname,rownum,columnnum,data):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	sheet.cell(rownum,columnnum).value = data
	workbook.save(file)
def Greenfill(file,sheetname,rownum,columnnum):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	green = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
	sheet.cell(rownum,columnnum).fill = green
	workbook.save(file)
def Redfill(file,sheetname,rownum,columnnum):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook[sheetname]
	red = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
	sheet.cell(rownum,columnnum).fill = red
	workbook.save(file)